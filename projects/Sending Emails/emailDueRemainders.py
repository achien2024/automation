import openpyxl
# works easier than google sheets API
# no permissions needed
# i coudln't get it to work
# for excel workbooks
import smtplib
# for sending emails
# requires IMAP server domain name
import os
# prob not needed unless you need to get the wkbk? 
import sys
# to quit the function if it drags on too long 

domain_library = {"GMAIL": 'imap.gmail.com', "OUTLOOK": 'imap-mail.outlook.com', "YAHOO": 'imap.mail.yahoo.com', 
                  "AT&T": 'imap.mail.att.net', "COMCAST": 'imap.comcast.net', "VERIZON": 'incoming.verizon.net'}

print("Which email domain do you use[Gmail, Outlook, Yahoo, AT&T, Comcast, Verizon]?")
domain = input().upper()

while domain not in domain_library.keys():
    print("Invalid domain. Please check for misspelling and characters.")
    domain = input().upper()

# log into server using SMTP
smtpObj = smtplib.SMTP(str(domain_library[domain]), 587)
# 587 is the port number to connect to SMTP server
smtpObj.ehlo()
# say hello to server
smtpObj.starttls()
# start the server 
print("Input your email address: ")
email_address = input()
print("Did you spell your address correctly (enter 'Y' to continue)?")
answer = input()
while answer != 'Y':
    print("Input your email address")
    email_address = input()
    print("Did you spell your address correctly (enter 'Y' to continue)?")
    answer = input()
print("Input API password: ")
password = input()
n = 0
max_attempts = 3

while n < max_attempts:
    try:
        smtpObj.login(str(email_address), str(password))
        break  # success, exit loop
    except smtplib.SMTPAuthenticationError:
        n += 1
        if n >= max_attempts:
            print("Too many failed attempts. Goodbye.")
            smtpObj.quit()
            sys.exit()
        else:
            print("Incorrect password. Try again:")
            password = input()

# load in the workbook
wb = openpyxl.load_workbook('EmailReminderDues.xlsx', data_only = True)
# data_only = True will ignore all formulas, only reads data from formula
members_sheet = wb['MEMBERS']
# grabs sheets from spreadsheets (MEMBERS)
email_index = 0
lastName_index = 1
firstName_index = 2
paid_index = 3
amount_index = 4
# column indexes
columns = list(members_sheet.iter_cols())
for i in range(1, members_sheet.max_row):
    if columns[paid_index][i].value == "no":
        first_name = list(columns[firstName_index])[i].value
        last_name = list(columns[lastName_index])[i].value
        amount = list(columns[amount_index])[i].value
        email = list(columns[email_index])[i].value

        subject = "Subject: [IMPORTANT] Missing Dues\n"
        body = f"""Hellp {first_name} {last_name},
                    
                   This is Aaron Chien. I am inquiring about missing dues of ${amount} Please pay before the end of this week or you will be dropped. Thank you!
                            
                   Best, Aaron
                """
        smtpObj.sendmail(str(email_address), str(email), subject + body)

        print(f"sending email to {email}")

smtpObj.quit()
