# TO DO
# set up excel workbook to calculate hours and calculate total payments
# check regex and test invalid inputs
# should I force end shift to be military time?
# comment on f""" and openpyxl methods to grab a column or row 

# libraries needed
import openpyxl
# indexing for rows and columns goes by 1 base
# for future rule, just do wb[sheet].cell(row = , column = ) # numeric based
# to grab a column, use wb[sheet]['COLUMN']
# to grab a row, use wb[sheet]['ROW']
import datetime
# when you run this program, automatically grabs the time
import re
# regex, for pattern matching

today = datetime.datetime.now()
year = today.year
month = today.month
day = today.day

wb = openpyxl.load_workbook('hourtracking.xlsx')
hours = wb['hour_tracking']
latest_row = hours.max_row + 1
pattern = re.compile(r'(\d?\d)(:)(\d\d)')
# parantheses are meant to represent groups
print("Input start shift: ")
start_shift = pattern.search(input())
n = 0
while not start_shift and n <= 3 or (int(start_shift.group(3)) > 60 or int(start_shift.group(1)) > 24):
    print("Wrong format, try again (HH:MM): ")
    start_shift = pattern.search(input())
    if n == 3:
        print("Too many tries. Goodbye")
        quit()

print("Input end shift: ")
end_shift = pattern.search(input())
n = 0 
while not end_shift and n <= 3 or (int(end_shift.group(3)) > 60 or int(start_shift.group(1)) > 24):
    print("Wrong format, try again (HH:MM): ")
    end_shift = pattern.search(input())
    if n == 3:
        print("Too many tries. Goodbye")
        quit()

hour_start, hour_end = start_shift.group(1), end_shift.group(1)
minute_start, minute_end = start_shift.group(3), end_shift.group(3)

start_datetime = datetime.datetime(year, month, day, hour = int(hour_start), minute = int(minute_start))
end_datetime = datetime.datetime(year, month, day, hour = int(hour_end), minute = int(minute_end))
total_time = round((end_datetime - start_datetime).total_seconds()/(60 * 60), 2)
hours.cell(row = latest_row, column = 1).value = f'''{today.month}/{today.day}/{today.year}'''
hours.cell(row = latest_row, column = 2).value = f'''{start_shift.group()}'''
hours.cell(row = latest_row, column = 3).value = f'''{end_shift.group()}'''
hours.cell(row = latest_row, column = 4).value = f'''{total_time}'''
wb.save('hourtracking.xlsx')

# Loop through each row (starting from row 2, assuming row 1 is headers)
total_wage = hours.cell(row = 3, column = 8)
calculate_wage = 0
for row in range(2, hours.max_row + 1):
    hours_cell = hours.cell(row=row, column=4)  # Total Hours

    # Skip if hours is empty
    if hours_cell.value is None:
        continue

    # Calculate and set total payment
    total_payment = round(float(hours_cell.value) * float(hours.cell(row = 2, column = 6).value), 2)
    calculate_wage += total_payment

total_wage.value = round(calculate_wage, 2)
wb.save('hourtracking.xlsx')
print(f"Wages calculated and updated successfully. Total: {total_wage.value}")

print("Thank you!")

